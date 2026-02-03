import json
import os
from openpyxl import load_workbook

def open_excel(path):
    if path.lower().endswith(".xlsm"):
        return load_workbook(path, keep_vba=True)
    else:
        return load_workbook(path)

DEVICES_XLSX = "devices.xlsx"


def add_device_to_excel(device, excel_path, excel_mapping):

    wb = open_excel(excel_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_index = {h: i for i, h in enumerate(headers)}

    # boş bir satır hazırla
    new_row = ["" for _ in headers]

    for field, header in excel_mapping.items():
        col_idx = header_index.get(header)
        if col_idx is not None:
            new_row[col_idx] = device.get(field, "")

    ws.append(new_row)
    wb.save(excel_path)





def load_devices_from_excel(path, mapping):
    devices = []

    # 1️⃣ ÖNCE PANDAS DENE
    try:
        import pandas as pd

        df = pd.read_excel(path)

        for _, row in df.iterrows():
            device = {}
            for field, header in mapping.items():
                device[field] = row.get(header)
            devices.append(device)

        return devices

    except Exception:
        pass   # pandas yoksa sessizce devam et

    # 2️⃣ FALLBACK → OPENPYXL

    wb = open_excel(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    header_index = {h: i for i, h in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        device = {}

        for field, header in mapping.items():
            idx = header_index.get(header)
            device[field] = row[idx] if idx is not None else None

        devices.append(device)

    return devices



def update_device_in_excel(old_ip, updated_device, excel_path, mapping):
    from openpyxl import load_workbook

    if not os.path.exists(excel_path):
        return

    wb = open_excel(excel_path)
    ws = wb.active

    # Excel başlıklarını al
    headers = [cell.value for cell in ws[1]]
    header_index = {h: i for i, h in enumerate(headers)}

    # IP hangi kolonda?
    ip_header = mapping.get("ip")
    if ip_header not in header_index:
        return

    ip_col = header_index[ip_header] + 1  # 1-based index

    for row in ws.iter_rows(min_row=2):
        if str(row[ip_col - 1].value) == str(old_ip):

            for field, header in mapping.items():
                col_idx = header_index.get(header)
                if col_idx is None:
                    continue

                row[col_idx].value = updated_device.get(field, "")

            break

    wb.save(excel_path)

def save_devices(devices):
    with open("devices.json", "w", encoding="utf-8") as f:
        json.dump(devices, f, indent=2, ensure_ascii=False)


def load_devices():
    if not os.path.exists("devices.json"):
        return []
    with open("devices.json", "r", encoding="utf-8") as f:
        return json.load(f)
    



def delete_device_from_excel(ip, excel_path, excel_mapping):
    if not excel_path or not excel_mapping:
        return

    wb = open_excel(excel_path)
    ws = wb.active

    ip_col_header = excel_mapping.get("ip")
    if not ip_col_header:
        return

    # IP kolon index
    ip_col_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == ip_col_header:
            ip_col_index = col
            break

    if not ip_col_index:
        return

    # IP eşleşen satırı bul ve sil
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=ip_col_index).value
        if str(cell_value).strip() == ip:
            ws.delete_rows(row)
            break

    wb.save(excel_path)