import tkinter as tk
from tkinter import ttk
import subprocess
import threading
import queue
import platform
import re
from datetime import datetime
from tkinter import messagebox
from device_loader import load_devices
from device_loader import load_devices_from_excel, save_devices
from tkinter import filedialog
import json
import os
import sys
import socket

# ---------------- UI THEME (Modern Dark) ----------------
# This block only affects styling (colors, rounded controls). Core logic is unchanged.

BG_COLOR = "#0E1117"
PANEL_COLOR = "#121826"
FG_COLOR = "#E6EAF2"
MUTED_FG = "#AAB3C5"
ACCENT = "#4C8DFF"
BORDER = "#3A4661"
SELECT_BG = "#2A3350"

def resource_path(relative_path: str) -> str:
    """Get absolute path to resource (works for PyInstaller onefile)."""
    base_path = getattr(sys, '_MEIPASS', os.path.abspath("."))
    return os.path.join(base_path, relative_path)

def load_ui_assets():
    """Load PNG assets used to fake rounded corners on Tk widgets."""
    assets = {}
    assets["bg"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "bg.png")))
    assets["btn_normal"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btn_normal.png")))
    assets["btn_hover"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btn_hover.png")))
    assets["btn_pressed"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btn_pressed.png")))
    assets["btn_disabled"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btn_disabled.png")))
    assets["btnw_normal"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btnw_normal.png")))
    assets["btnw_hover"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btnw_hover.png")))
    assets["btnw_pressed"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btnw_pressed.png")))
    assets["btnw_disabled"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "btnw_disabled.png")))
    assets["entry_bg"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "entry_bg.png")))
    assets["search_bg"] = tk.PhotoImage(file=resource_path(os.path.join("assets", "search_bg.png")))
    return assets

def apply_ttk_dark_style(root):
    """Style ttk widgets (Treeview + Scrollbars) for a modern dark look."""
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except:
        pass

    style.configure(".", background=BG_COLOR, foreground=FG_COLOR, fieldbackground=PANEL_COLOR)
    style.configure("TFrame", background=BG_COLOR)
    style.configure("TLabel", background=BG_COLOR, foreground=FG_COLOR)

    # Scrollbar
    style.configure("Vertical.TScrollbar", background=BG_COLOR, troughcolor=BG_COLOR, bordercolor=BG_COLOR, arrowcolor=FG_COLOR)
    style.configure("Horizontal.TScrollbar", background=BG_COLOR, troughcolor=BG_COLOR, bordercolor=BG_COLOR, arrowcolor=FG_COLOR)

    # Treeview (table)
    style.configure(
        "Treeview",
        background=PANEL_COLOR,
        fieldbackground=PANEL_COLOR,
        foreground=FG_COLOR,
        borderwidth=0,
        relief="flat",
        rowheight=26
    )
    style.map(
        "Treeview",
        background=[("selected", SELECT_BG)],
        foreground=[("selected", "#FFFFFF")]
    )

    style.configure(
        "Treeview.Heading",
        background="#1B2130",
        foreground="#FFFFFF",
        relief="flat"
    )
    style.map("Treeview.Heading", background=[("active", "#25304A")])

        # ===== EXCEL MAPPING COMBOBOX (Windows + macOS) =====
    style.configure(
        "ExcelMap.TCombobox",
        foreground="#000000",
        fieldbackground="#FFFFFF",
        background="#FFFFFF"
    )

    style.map(
        "ExcelMap.TCombobox",
        foreground=[
            ("readonly", "#000000"),
            ("!disabled", "#000000")
        ],
        fieldbackground=[
            ("readonly", "#FFFFFF"),
            ("!disabled", "#FFFFFF")
        ]
    )

    return style

def make_rounded_entry(parent, bg_image, *, font, inner_pad=10):
    """Create a rounded-looking Entry by placing it on top of a PNG."""
    w = bg_image.width()
    h = bg_image.height()
    c = tk.Canvas(parent, width=w, height=h, bg=BG_COLOR, highlightthickness=0, bd=0)
    c.create_image(0, 0, anchor="nw", image=bg_image)

    e = tk.Entry(
        c,
        bd=0,
        relief="flat",
        highlightthickness=0,
        bg=PANEL_COLOR,
        fg=FG_COLOR,
        insertbackground=FG_COLOR,
        font=font,
    )

    # Place the entry on top of the image with padding.
    c.create_window(
        inner_pad,
        h // 2,
        anchor="w",
        window=e,
        width=max(10, w - inner_pad * 2),
        height=max(10, h - 6),
    )
    return c, e

def style_rounded_button(btn, assets, *, wide=False):
    """Apply rounded PNG background + hover/pressed behavior to a tk.Button.

    NOTE: We do NOT rely on Tk's DISABLED state, because Tk will gray-out image buttons
    with a very "old" look. Instead we keep the widget state NORMAL and implement an
    internal enabled flag + a dedicated disabled image.
    """
    normal = assets["btnw_normal"] if wide else assets["btn_normal"]
    hover = assets["btnw_hover"] if wide else assets["btn_hover"]
    pressed = assets["btnw_pressed"] if wide else assets["btn_pressed"]
    disabled = assets["btnw_disabled"] if wide else assets["btn_disabled"]

    btn.configure(
        image=normal,
        compound="center",
        bd=0,
        relief="flat",
        highlightthickness=0,
        bg=BG_COLOR,
        fg=FG_COLOR,
        activeforeground=FG_COLOR,
        activebackground=BG_COLOR,
        cursor="hand2"
    )

    # Internal UI state
    btn._ui_imgs = (normal, hover, pressed, disabled)
    btn._ui_enabled = True
    btn._ui_pressed = False

    def _set_image(img):
        try:
            btn.configure(image=img)
        except tk.TclError:
            # widget may be destroyed during shutdown
            return

    def ui_set_enabled(enabled: bool):
        btn._ui_enabled = bool(enabled)
        if btn._ui_enabled:
            btn.configure(cursor="hand2", fg=FG_COLOR)
            _set_image(normal)
        else:
            # keep the modern look (no default gray overlay)
            btn.configure(cursor="arrow", fg=MUTED_FG)
            _set_image(disabled)

    btn.ui_set_enabled = ui_set_enabled  # attach helper

    def on_enter(_):
        if not btn._ui_enabled:
            return
        if not btn._ui_pressed:
            _set_image(hover)

    def on_leave(_):
        if not btn._ui_enabled:
            return
        btn._ui_pressed = False
        _set_image(normal)

    def on_press(_):
        if not btn._ui_enabled:
            return "break"
        btn._ui_pressed = True
        _set_image(pressed)
        return "break"

    def on_release(event):
        # Stop Tk's default button invoke; we'll invoke ourselves if enabled.
        if not btn._ui_enabled:
            btn._ui_pressed = False
            _set_image(disabled)
            return "break"

        btn._ui_pressed = False

        # Only invoke if mouse is still inside the widget bounds
        x, y = event.x, event.y
        if 0 <= x < btn.winfo_width() and 0 <= y < btn.winfo_height():
            _set_image(hover)
            try:
                btn.invoke()
            except tk.TclError:
                pass
        else:
            _set_image(normal)
        return "break"

    # Bindings (break prevents old-school gray state behavior)
    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    btn.bind("<ButtonPress-1>", on_press)
    btn.bind("<ButtonRelease-1>", on_release)





# ---------------- PLATFORM ----------------ge=hover)

    btn.bind("<Enter>", on_enter)
    btn.bind("<Leave>", on_leave)
    btn.bind("<ButtonPress-1>", on_press)
    btn.bind("<ButtonRelease-1>", on_release)





# ---------------- PLATFORM ----------------
IS_WINDOWS = platform.system().lower() == "windows"
FONT_NAME = "Segoe UI" if IS_WINDOWS else "Helvetica"

# ---------------- GLOBAL STATE ----------------
CONFIG_FILE = "config.json"
current_task = None   # None / "PING" / "TRACE" / "NSLOOKUP" / "PORTTEST" / "BULK"
open_ports_found = []
excel_path = None
excel_mapping = None
bulk_total = 0
bulk_done = 0
PAGE_SIZE = 100
current_page = 1
total_pages = 1
traceroute_process = None
nslookup_process = None
REQUIRED_FIELDS = [
    "ip",
    "device",
    "name",
    "model",
    "mac",
    "location",
    "unit",
    "description"
]
COLUMN_WIDTHS = {
    "Cihaz": 150,
    "IP": 130,
    "Ping (ms)": 100,
    "Son Ping": 170,
    "Device Name": 160,   # 👈 EKLE
    "Model": 120,
    "MAC": 160,
    "Location": 140,
    "Unit": 120,
    "Description": 200
}
current_sort_column = None
current_sort_reverse = False

devices = []
current_ip = None
is_running = False
ping_process = None
is_bulk_running = False
ping_thread = None
ui_queue = queue.Queue()
started_from_entry = False
is_port_test_running = False
stop_port_test_flag = False

def update_column_headers():
    for col, field in COLUMN_TO_FIELD.items():
        text = col

        # 🔍 Filtre aktif mi?
        if active_filters.get(field):
            text += " 🔍"

        # ⬆⬇ sıralama oku
        if col == current_sort_column:
            text += " ▼" if current_sort_reverse else " ▲"
        else:
            text += " ▼"

        device_tree.heading(col, text=text)

def update_tree_item_for_ip(ip):
    for item in device_tree.get_children():
        values = device_tree.item(item)["values"]
        if values and values[1] == ip:
            device = next((d for d in devices if d["ip"] == ip), None)
            if not device:
                return

            latency_txt = "-" if device.get("latency") is None else f"{device['latency']:.1f}"

            device_tree.item(
                item,
                values=(
                    device.get("device", ""),
                    device.get("ip", ""),
                    latency_txt,
                    device.get("last_ping") or "-",
                    device.get("name", ""),   # 👈 EKLE
                    device.get("model", ""),
                    device.get("mac", ""),
                    device.get("location", ""),
                    device.get("unit", ""),
                    device.get("description", "")
                ),
                tags=(device.get("status", "UNKNOWN"),)
            )
            return

# ---------------- PING HELPERS ----------------
def nslookup_command(target, dns_server=None):
    # target = domain veya ip olabilir
    if dns_server:
        return ["nslookup", target, dns_server]
    return ["nslookup", target]

def nslookup_worker(target, dns_server=None):
    global nslookup_process
    flags = subprocess.CREATE_NO_WINDOW if IS_WINDOWS else 0

    ui_queue.put(("NSLOOKUP_START", target, None))

    nslookup_process = subprocess.Popen(
        nslookup_command(target, dns_server),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        creationflags=flags
    )

    for line in nslookup_process.stdout:
        ui_queue.put(("NSLOOKUP", target, line))

    ui_queue.put(("NSLOOKUP_DONE", target, None))
    nslookup_process = None

def extract_ping_ms(text):
    text = text.lower()

    # time=14.2 ms | time<1ms | time=1ms
    match = re.search(r"time[=<]?\s*([\d\.]+)\s*ms", text)
    if match:
        return float(match.group(1))

    # bazı sistemlerde boşluk yok: time<1ms
    match = re.search(r"time[=<]?([\d\.]+)ms", text)
    if match:
        return float(match.group(1))

    return None

def single_ping(ip, timeout=2):
    try:
        if IS_WINDOWS:
            cmd = ["ping", "-n", "1", "-w", str(timeout * 1000), ip]
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW

            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                startupinfo=startupinfo,
                creationflags=subprocess.CREATE_NO_WINDOW
            )
        else:
            cmd = ["ping", "-c", "1", ip]
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True
            )

        output, _ = proc.communicate(timeout=timeout + 1)
        return extract_ping_ms(output)

    except Exception:
        return None
    
from concurrent.futures import ThreadPoolExecutor, as_completed

COMMON_PORTS = [
    (22, "SSH"),
    (80, "HTTP"),
    (443, "HTTPS"),
    (3389, "RDP"),
]
PRIORITY_PORTS = [
    (22,  "SSH"),
    (80,  "HTTP"),
    (443, "HTTPS"),
    (445, "SMB"),
    (3389, "RDP"),
    (53,  "DNS"),
    (139, "NETBIOS"),
    (161, "SNMP"),
    (8080, "HTTP-ALT"),
]

SECONDARY_PORTS = [
    # ===== File Transfer / Legacy =====
    (20,  "FTP-DATA"),
    (21,  "FTP"),
    (22,  "SSH"),           # priority'de var ama tekrar sorun değil
    (23,  "TELNET"),
    (69,  "TFTP"),
    (115, "SFTP"),
    (989, "FTPS-DATA"),
    (990, "FTPS"),
    (873, "RSYNC"),

    # ===== Web / Proxy =====
    (80,   "HTTP"),
    (81,   "HTTP-ALT"),
    (443,  "HTTPS"),
    (8000, "HTTP-ALT"),
    (8008, "HTTP-ALT"),
    (8080, "HTTP-PROXY"),
    (8081, "HTTP-ALT"),
    (8088, "HTTP-ALT"),
    (8443, "HTTPS-ALT"),
    (8888, "HTTP-ALT"),

    # ===== DNS / DHCP / Time =====
    (53,   "DNS"),
    (67,   "DHCP-SRV"),
    (68,   "DHCP-CLI"),
    (123,  "NTP"),

    # ===== Mail =====
    (25,   "SMTP"),
    (110,  "POP3"),
    (143,  "IMAP"),
    (465,  "SMTPS"),
    (587,  "SMTP-SUB"),
    (993,  "IMAPS"),
    (995,  "POP3S"),

    # ===== Windows / AD / SMB =====
    (88,   "KERBEROS"),
    (135,  "RPC-EPMAP"),
    (137,  "NETBIOS-NS"),
    (138,  "NETBIOS-DGM"),
    (139,  "NETBIOS-SSN"),
    (389,  "LDAP"),
    (445,  "SMB"),
    (464,  "KPASSWD"),
    (636,  "LDAPS"),
    (3268, "GC-LDAP"),
    (3269, "GC-LDAPS"),

    # ===== Remote Access =====
    (5900, "VNC"),
    (5901, "VNC-ALT"),
    (5938, "TeamViewer"),
    (3389, "RDP"),
    (22,   "SSH"),
    (2222, "SSH-ALT"),
    (5800, "VNC-WEB"),

    # ===== Printing =====
    (515,  "LPD"),
    (631,  "IPP"),

    # ===== Network Management =====
    (161,  "SNMP"),
    (162,  "SNMP-TRAP"),
    (514,  "SYSLOG"),
    (179,  "BGP"),

    # ===== Databases =====
    (1433, "MSSQL"),
    (1434, "MSSQL-BROWSER"),
    (1521, "ORACLE"),
    (2049, "NFS"),
    (27017, "MongoDB"),
    (3306, "MySQL"),
    (5432, "PostgreSQL"),
    (6379, "Redis"),
    (11211, "Memcached"),
    (9200, "Elastic"),
    (9300, "Elastic-Transport"),

    # ===== Message Brokers / IoT =====
    (1883,  "MQTT"),
    (8883,  "MQTT-TLS"),
    (5672,  "AMQP"),
    (15672, "RabbitMQ-UI"),
    (9092,  "Kafka"),
    (9093,  "Kafka-SSL"),

    # ===== VoIP =====
    (5060, "SIP"),
    (5061, "SIPS"),
    (1720, "H.323"),

    # ===== VPN / Tunneling =====
    (500,  "IKE"),
    (1701, "L2TP"),
    (1723, "PPTP"),
    (4500, "IPSEC-NAT-T"),
    (1194, "OpenVPN"),
    (51820, "WireGuard"),

    # ===== Virtualization / Remote Mgmt =====
    (902,  "VMware"),
    (903,  "VMware-ALT"),
    (9443, "vSphere-ALT"),
    (5985, "WinRM-HTTP"),
    (5986, "WinRM-HTTPS"),

    # ===== DevOps / Containers =====
    (2375, "Docker"),
    (2376, "Docker-TLS"),
    (6443, "Kubernetes-API"),

    # ===== Git / CI =====
    (9418, "GIT"),

    # ===== Media / Streaming =====
    (554,  "RTSP"),
    (1935, "RTMP"),

    # ===== Other Common =====
    (111,  "RPC"),
    (2048, "DLS-MON"),
    (6667, "IRC"),
]
# ✅ FULL PORT SCAN (1-65535)
ALL_PORTS = [(p, f"PORT-{p}") for p in range(1, 65536)]

def check_port(ip, port, timeout=1.0):
    try:
        with socket.create_connection((ip, port), timeout=timeout):
            return True
    except:
        return False
   
def port_test_worker(ip, mode="fast"):
    global is_port_test_running, stop_port_test_flag, open_ports_found
    open_ports_found = []   # ✅ her testte sıfırla
    """
    mode = "fast"  -> PRIORITY + SECONDARY
    mode = "full"  -> PRIORITY + ALL_PORTS
    """
    is_port_test_running = True
    stop_port_test_flag = False

    if mode == "full":
        phase2_ports = ALL_PORTS
        phase2_name = "Tüm Portlar (1-65535)"
        phase2_timeout = 0.35
        phase2_workers = 80
    else:
        phase2_ports = SECONDARY_PORTS
        phase2_name = "Diğer Portlar"
        phase2_timeout = 0.9
        phase2_workers = 30

    total_ports = len(PRIORITY_PORTS) + len(phase2_ports)
    done = 0
    open_count = 0
    closed_count = 0

    ui_queue.put(("PORT_TEST_START", ip, total_ports))

    def run_phase(ports, phase_name, timeout, workers):
        nonlocal done, open_count, closed_count

        ui_queue.put(("PORT_TEST_PHASE", ip, phase_name))

        def check_one(p):
            port, name = p
            ok = check_port(ip, port, timeout=timeout)
            return (port, name, ok)

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = [executor.submit(check_one, p) for p in ports]

            for f in as_completed(futures):
                if stop_port_test_flag:
                    return False

                port, name, ok = f.result()

                done += 1
                if ok:
                    open_count += 1
                else:
                    closed_count += 1

                ui_queue.put(("PORT_TEST_RESULT", ip, (port, name, ok, phase_name)))
                ui_queue.put(("PORT_TEST_PROGRESS", ip, (done, total_ports, open_count, closed_count)))

        return True

    ok1 = run_phase(PRIORITY_PORTS, "Öncelikli Portlar", timeout=0.6, workers=40)
    if not ok1:
        ui_queue.put(("PORT_TEST_CANCELLED", ip, (done, total_ports, open_count, closed_count)))
        is_port_test_running = False
        return

    ok2 = run_phase(phase2_ports, phase2_name, timeout=phase2_timeout, workers=phase2_workers)
    if not ok2:
        ui_queue.put(("PORT_TEST_CANCELLED", ip, (done, total_ports, open_count, closed_count)))
        is_port_test_running = False
        return

    ui_queue.put(("PORT_TEST_DONE", ip, (done, total_ports, open_count, closed_count)))
    is_port_test_running = False



def ip_exists(ip, exclude_device=None):
    for d in devices:
        if d["ip"] == ip:
            if exclude_device and d is exclude_device:
                continue
            return True
    return False

def bulk_ping_worker(devices_to_ping):
    global is_bulk_running

    def ping_one(device):
        ip = device["ip"]
        ms = single_ping(ip)
        ui_queue.put(("BULK", ip, ms))

    with ThreadPoolExecutor(max_workers=10) as executor:
        executor.map(ping_one, devices_to_ping)

    ui_queue.put(("BULK_DONE", None, None))

def device_matches_filters(device):
    # 1️⃣ Checkbox filtreleri
    for field, selected_values in active_filters.items():
        if selected_values:
            if device.get(field) not in selected_values:
                return False

    # 2️⃣ Global arama
    if search_text:
        haystack = " ".join(
            str(device.get(k, "")).lower()
            for k in [
                "device", "ip", "model",
                "mac", "location", "unit", "description"
            ]
        )
        if search_text.lower() not in haystack:
            return False

    return True

def status_by_latency(ms):
    if ms is None:
        return "DOWN"
    if ms < 50:
        return "FAST"
    elif ms < 100:
        return "NORMAL"
    elif ms < 200:
        return "SLOW"
    return "VERY_SLOW"


def sort_devices_by_column(col, reverse=False):
    global current_page
    current_page = 1
    global current_sort_column, current_sort_reverse

    if col == "IP":
        devices.sort(key=lambda d: ip_to_tuple(d.get("ip", "")), reverse=reverse)
    else:
        field = COLUMN_TO_FIELD.get(col)
        devices.sort(
            key=lambda d: (d.get(field) or "").lower(),
            reverse=reverse
        )

    current_sort_column = col
    current_sort_reverse = reverse

    refresh_device_list(keep_selection=True)
    update_column_headers()

def ping_command(ip):
    return ["ping", "-t", ip] if IS_WINDOWS else ["ping", ip]

def traceroute_command(ip):
    # -d = DNS çözümleme kapalı (daha hızlı)
    if IS_WINDOWS:
        return ["tracert", "-d", ip]
    else:
        return ["traceroute", ip]


def traceroute_worker(ip):
    global traceroute_process
    flags = subprocess.CREATE_NO_WINDOW if IS_WINDOWS else 0

    traceroute_process = subprocess.Popen(
        traceroute_command(ip),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        creationflags=flags
    )

    for line in traceroute_process.stdout:
        ui_queue.put(("TRACE", ip, line))

    ui_queue.put(("TRACE_DONE", ip, None))
    traceroute_process = None


def ip_to_tuple(ip):
    try:
        return tuple(int(x) for x in ip.split("."))
    except:
        return (0, 0, 0, 0)

def load_config():
    global excel_path, excel_mapping

    if not os.path.exists(CONFIG_FILE):
        return

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    excel_path = data.get("excel_path")
    excel_mapping = data.get("excel_mapping")

    if not excel_path or not os.path.exists(excel_path):
        excel_path = None
        excel_mapping = None

def save_config():
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(
            {
                "excel_path": excel_path,
                "excel_mapping": excel_mapping
            },
            f,
            ensure_ascii=False,
            indent=2
        )

# ---------------- FILTER STATE ----------------
FILTERABLE_FIELDS = {
    "device": "Cihaz",
    "ip": "IP",
    "name": "Device Name",   # 👈 EKLE
    "model": "Model",
    "mac": "MAC",
    "location": "Location",
    "unit": "Unit",
    "description": "Description"
}
COLUMN_TO_FIELD = {
    "Cihaz": "device",
    "IP": "ip",
    "Device Name": "name",   # 👈 EKLE
    "Model": "model",
    "MAC": "mac",
    "Location": "location",
    "Unit": "unit",
    "Description": "description"
}
# 🔹 Treeview
cols = (
    "Cihaz",
    "IP",
    "Ping (ms)",
    "Son Ping",
    "Device Name",   # 👈 YENİ
    "Model",
    "MAC",
    "Location",
    "Unit",
    "Description"
)

active_filters = {key: set() for key in FILTERABLE_FIELDS}
search_text = ""

# ---------------- PING LOOP ----------------
def ping_loop(ip):
    global ping_process

    flags = subprocess.CREATE_NO_WINDOW if IS_WINDOWS else 0

    ping_process = subprocess.Popen(
        ping_command(ip),
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        creationflags=flags
    )

    for line in ping_process.stdout:
        if not is_running:
            break
        ui_queue.put(("SINGLE", ip, line))

    try:
        ping_process.terminate()
    except Exception:
        pass

# ---------------- UI QUEUE ----------------

def process_ui_queue():
    MAX_ITEMS_PER_TICK = 200
    count = 0

    while not ui_queue.empty() and count < MAX_ITEMS_PER_TICK:
        item = ui_queue.get()
        count += 1

        item_type, ip, payload = item
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 🔵 TEKLİ PING
        if item_type == "SINGLE":
            line = payload

            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, line)
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

            ms = extract_ping_ms(line)

            for d in devices:
                if d["ip"] == ip:
                    d["latency"] = ms
                    d["last_ping"] = now
                    d["status"] = status_by_latency(ms)
                    update_tree_item_for_ip(ip)
                    break

            save_devices(devices)

        # 🟢 TOPLU PING
        elif item_type == "BULK":
            global bulk_done

            bulk_done += 1   # 🔥 BAŞARILI / BAŞARISIZ FARK ETMEZ

            bulk_status_label.config(
                text=f"Toplu Ping: {bulk_done} / {bulk_total}"
            )

            ms = payload  # None olabilir, sorun değil

            for d in devices:
                if d["ip"] == ip:
                    d["latency"] = ms
                    d["last_ping"] = now
                    d["status"] = status_by_latency(ms)
                    update_tree_item_for_ip(ip)
                    break

                # 🧭 TRACEROUTE
        elif item_type == "TRACE":
            line = payload

            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, line)
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

        elif item_type == "TRACE_DONE":
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, "\n--- Traceroute tamamlandı ---\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

            unlock_ui()

                # 🔌 PORT TEST
        elif item_type == "PORT_TEST_START":
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, f"\n--- Port Test Başladı: {ip} ---\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

        elif item_type == "PORT_TEST_PHASE":
            phase_name = payload
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, f"\n>>> {phase_name} <<<\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

        elif item_type == "PORT_TEST_RESULT":
            port, name, ok, phase_name = payload

            # sadece OPEN'ları kaydet
            if ok:
                open_ports_found.append((port, name))

                # OPEN olunca ekrana da yaz (az satır -> UI donmaz)
                output_box.config(state=tk.NORMAL)
                output_box.insert(tk.END, f"{port:<5} ({name}) -> OPEN ✅\n")
                output_box.see(tk.END)
                output_box.config(state=tk.DISABLED)

        elif item_type == "PORT_TEST_DONE":
            done, total, open_count, closed_count = payload

            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, f"\n--- Port Test Bitti: {ip} ---\n")
            output_box.insert(tk.END, f"Toplam: {total} | OPEN: {open_count} | CLOSED: {closed_count}\n\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

            bulk_status_label.config(text="")
            start_btn.config(text="▶ Başlat")
            bulk_status_label.config(text="")
            # ✅ Port test bitti -> OPEN portları tek seferde yaz
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, "\n=== OPEN PORTLAR ===\n")

            for port, name in sorted(open_ports_found):
                output_box.insert(tk.END, f"{port:<5} ({name}) -> OPEN ✅\n")

            output_box.insert(tk.END, "\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)
            unlock_ui()
        elif item_type == "PORT_TEST_CANCELLED":
            done, total, open_count, closed_count = payload

            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, f"\n--- Port Test DURDURULDU: {ip} ---\n")
            output_box.insert(tk.END, f"İlerleme: {done}/{total} | OPEN: {open_count} | CLOSED: {closed_count}\n\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)
            unlock_ui()

            bulk_status_label.config(text="Port test durduruldu.")
            root.after(3000, lambda: bulk_status_label.config(text=""))

            start_btn.config(text="▶ Başlat")

        elif item_type == "PORT_TEST_PROGRESS":
            done, total, open_count, closed_count = payload
            bulk_status_label.config(
                text=f"Port Test: {done}/{total}  |  OPEN:{open_count}  CLOSED:{closed_count}"
            )        

                # 🌐 NSLOOKUP
        elif item_type == "NSLOOKUP_START":
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, f"\n--- NSLOOKUP Başladı: {ip} ---\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

        elif item_type == "NSLOOKUP":
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, payload)
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)

        elif item_type == "NSLOOKUP_DONE":
            output_box.config(state=tk.NORMAL)
            output_box.insert(tk.END, f"--- NSLOOKUP Bitti: {ip} ---\n\n")
            output_box.see(tk.END)
            output_box.config(state=tk.DISABLED)
            unlock_ui()

        # 🔴 TOPLU PING BİTTİ
        elif item_type == "BULK_DONE":
            global is_bulk_running

            is_bulk_running = False
            save_devices(devices)
            start_btn.ui_set_enabled(True)
            refresh_btn.ui_set_enabled(True)
            add_btn.ui_set_enabled(True)
            bulk_status_label.config(
            text=f"Toplu Ping tamamlandı ({bulk_total} / {bulk_total})"
            )
              # ⏱ 3 saniye sonra temizle
            root.after(5000, lambda: bulk_status_label.config(text=""))
            unlock_ui()
  

    root.after(30, process_ui_queue)

# ---------------- ACTIONS ----------------
def lock_ui(task_name):
    global current_task
    current_task = task_name

    refresh_btn.ui_set_enabled(False)
    add_btn.ui_set_enabled(False)
    excel_btn.ui_set_enabled(False)

    start_btn.config(text="⏹ Durdur")
    start_btn.ui_set_enabled(True)

def unlock_ui():
    global current_task
    current_task = None

    refresh_btn.ui_set_enabled(True)
    add_btn.ui_set_enabled(True)
    excel_btn.ui_set_enabled(True)

    start_btn.config(text="▶ Başlat")
    start_btn.ui_set_enabled(True)

def can_start_new_task():
    if current_task is not None:
        messagebox.showwarning("Meşgul", "Şu an bir işlem çalışıyor. Önce DURDUR.")
        return False
    return True

def start_ping(event=None):
    global is_running, current_ip, ping_thread, started_from_entry
    bulk_status_label.config(text="")
    if not can_start_new_task():
        return
    lock_ui("PING")
    ip = ip_entry.get().strip()
    if not ip:
        return

    # 🔴 BU SATIR KRİTİK
    started_from_entry = True

    stop_ping_silent()

    is_running = True
    current_ip = ip
    start_btn.config(text="⏹ Durdur")

    output_box.config(state=tk.NORMAL)
    output_box.delete("1.0", tk.END)
    output_box.config(state=tk.DISABLED)

    while not ui_queue.empty():
        ui_queue.get_nowait()

    ping_thread = threading.Thread(target=ping_loop, args=(ip,), daemon=True)
    ping_thread.start()

def start_ping_from_menu():
    global started_from_entry
    started_from_entry = False
    start_ping()

def start_traceroute_selected():
    if not can_start_new_task():
        return
    lock_ui("TRACE")
    global started_from_entry

    # seçili cihaz yoksa entry'den IP al
    sel = device_tree.selection()
    if sel:
        ip = device_tree.item(sel[0])["values"][1]
    else:
        ip = ip_entry.get().strip()

    if not ip:
        messagebox.showwarning("Uyarı", "Traceroute için bir IP seçin veya girin.")
        return

    # output'u temizle
    output_box.config(state=tk.NORMAL)
    output_box.delete("1.0", tk.END)
    output_box.config(state=tk.DISABLED)

    # ping çalışıyorsa durdur (aynı anda karışmasın)
    stop_ping()

    # traceroute thread
    threading.Thread(target=traceroute_worker, args=(ip,), daemon=True).start()

def start_port_test_selected(mode="fast"):
    if not can_start_new_task():
        return
    lock_ui("PORTTEST")
    # seçili cihaz varsa oradan al
    sel = device_tree.selection()
    if sel:
        ip = device_tree.item(sel[0])["values"][1]
    else:
        ip = ip_entry.get().strip()

    if not ip:
        messagebox.showwarning("Uyarı", "Port testi için bir IP seçin veya girin.")
        return

    # output'u temizle
    output_box.config(state=tk.NORMAL)
    output_box.delete("1.0", tk.END)
    output_box.config(state=tk.DISABLED)

    # ping çalışıyorsa durdur
    stop_ping()

    global is_running
    is_running = False

    # Start butonu port test sırasında "Durdur" gibi dursun
    start_btn.config(text="⏹ Durdur")
    start_btn.ui_set_enabled(True)

    # ✅ mode'u burada gönderiyoruz
    threading.Thread(target=port_test_worker, args=(ip, mode), daemon=True).start()

def stop_port_test():
    global stop_port_test_flag, is_port_test_running
    stop_port_test_flag = True

    # UI'yi anında toparla
    start_btn.config(text="▶ Başlat")
    bulk_status_label.config(text="Port test durduruluyor...")
    root.after(2000, lambda: bulk_status_label.config(text=""))

def start_nslookup_selected():
    if not can_start_new_task():
        return
    lock_ui("NSLOOKUP")

    # seçili cihaz varsa IP'sini al
    sel = device_tree.selection()
    if sel:
        target = device_tree.item(sel[0])["values"][1]
    else:
        target = ip_entry.get().strip()

    if not target:
        messagebox.showwarning("Uyarı", "NSLOOKUP için bir hedef girin (IP veya domain).")
        return

    output_box.config(state=tk.NORMAL)
    output_box.delete("1.0", tk.END)
    output_box.config(state=tk.DISABLED)

    stop_ping()

    threading.Thread(target=nslookup_worker, args=(target,), daemon=True).start()

def stop_ping_silent():
    global is_running, ping_process
    is_running = False

    if ping_process:
        try:
            ping_process.terminate()
        except:
            pass
        ping_process = None

def stop_ping(event=None):
    global is_running, ping_process

    is_running = False
    start_btn.config(text="▶ Başlat")
        # Ping durdu, artık entry'yi otomatik doldurabiliriz
    global started_from_entry

    if ping_process:
        try:
            ping_process.terminate()
        except Exception:
            pass
        ping_process = None
    
    unlock_ui()

def stop_traceroute():
    global traceroute_process

    if traceroute_process:
        try:
            traceroute_process.terminate()
        except:
            pass
        traceroute_process = None

    unlock_ui()

def stop_nslookup():
    global nslookup_process

    if nslookup_process:
        try:
            nslookup_process.terminate()
        except:
            pass
        nslookup_process = None

    unlock_ui()


def toggle_ping():
    global current_task

    # Port Test çalışıyorsa Durdur butonu onu iptal etsin
    if current_task == "PORTTEST":
        stop_port_test()
        return

    # Traceroute / NSLOOKUP durdurma yok (istersen ekleriz ama şimdilik engelleme yeter)
    # Ping çalışıyorsa
    if current_task == "PING":
        stop_ping()
        return

    # Bulk çalışıyorsa şimdilik durdurma yok (istersen ekleriz)
    if current_task == "BULK":
        messagebox.showinfo("Bilgi", "Toplu Ping iptali şu an yok.")
        return
    
    if current_task == "TRACE":
        stop_traceroute()
        return

    if current_task == "NSLOOKUP":
        stop_nslookup()
        return
    # hiç iş yoksa ping başlat
    start_ping()


def refresh_from_excel():
    global devices

    if not excel_path or not excel_mapping:
        messagebox.showwarning(
            "Excel",
            "Excel seçilmemiş veya kolon eşleştirmesi yok."
        )
        return

    # 1️⃣ Excel'den cihazları oku
    excel_devices = load_devices_from_excel(excel_path, excel_mapping)

    # 2️⃣ Eski ping bilgilerini koru
    new_devices = []

    for ex in excel_devices:
        old = next((d for d in devices if d.get("ip") == ex.get("ip")), None)

        if old:
            ex["latency"] = old.get("latency")
            ex["last_ping"] = old.get("last_ping")
            ex["status"] = old.get("status", "UNKNOWN")
        else:
            ex["latency"] = None
            ex["last_ping"] = None
            ex["status"] = "UNKNOWN"

        new_devices.append(ex)

    # 3️⃣ RAM + JSON güncelle
    devices = new_devices
    save_devices(devices)

    # 4️⃣ Listeyi yenile
    refresh_device_list(keep_selection=True)

def open_mapping_window(excel_headers, on_done):
    messagebox.showinfo(
    "DEBUG",
    "Excel seçildi.\nKolon eşleştirme penceresi açılıyor."
)
    win = tk.Toplevel(root)

    # 🔥 EXE + WINDOWS İÇİN ZORUNLU
    win.withdraw()          # önce gizle
    win.transient(root)
    win.grab_set()
    win.lift()
    win.focus_force()

    win.title("Excel Kolon Eşleştirme")
    win.geometry("520x500")

    # 🔥 sonra göster
    win.deiconify()

    tk.Label(
        win,
        text="Excel kolonlarını uygulama alanlarıyla eşleştir",
        font=(FONT_NAME, 11, "bold")
    ).pack(pady=10)

    mapping_vars = {}

    frame = tk.Frame(win)
    frame.pack(padx=10, pady=10)

    for header in excel_headers:
        row = tk.Frame(frame)
        row.pack(fill=tk.X, pady=3)

        tk.Label(row, text=header, width=25, anchor="w").pack(side=tk.LEFT)

        var = tk.StringVar(value="")
        combo = ttk.Combobox(
            row,
            textvariable=var,
            values=[""] + REQUIRED_FIELDS,
            state="normal",                 # ⬅️ readonly YOK
            width=20,
            style="ExcelMap.TCombobox"      # ⬅️ bizim stil
        )

        # yazı yazmayı kapat (readonly gibi davransın)
        combo.bind("<Key>", lambda e: "break")
        combo.pack(side=tk.LEFT)

        mapping_vars[header] = var

    def apply():
        mapping = {}
        for header, var in mapping_vars.items():
            if var.get():
                mapping[var.get()] = header

        if "ip" not in mapping:
            messagebox.showerror("Hata", "IP alanı mutlaka eşleştirilmeli")
            return

        win.destroy()
        on_done(mapping)

    tk.Button(win, text="Tamam", command=apply).pack(pady=10)

def select_excel_file():
    global excel_path

    path = filedialog.askopenfilename(
        title="Excel dosyasını seç",
        filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")]
    )

    if not path:
        return
    
    if not path.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        messagebox.showerror(
            "Geçersiz Dosya",
            "Lütfen .xlsx formatında bir Excel dosyası seçin"
        )
        return

    excel_path = path
    save_config()

    # 🔴 DEBUG: buraya geliyor mu?
    messagebox.showinfo("DEBUG", "Excel dosyası seçildi")

    try:
        from openpyxl import load_workbook

        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]

    except Exception as e:
        messagebox.showerror(
            "Excel Okuma Hatası",
            f"Excel okunamadı:\n\n{e}"
        )
        return

    messagebox.showinfo("DEBUG", "Excel okundu, eşleştirme açılıyor")

    def on_mapping_done(mapping):
        global excel_mapping, devices
        excel_mapping = mapping
        save_config()

        devices = load_devices_from_excel(excel_path, excel_mapping)
        refresh_device_list()

    root.after(
        0,
        lambda: open_mapping_window(headers, on_mapping_done))
# ---------------- DEVICE LIST ----------------
def extend_selection(direction):
    items = device_tree.get_children()
    if not items:
        return "break"

    sel = device_tree.selection()

    # hiç seçim yoksa → ilk satırı seç
    if not sel:
        device_tree.selection_set(items[0])
        device_tree.focus(items[0])
        return "break"

    # son seçili satır
    last = sel[-1]
    index = items.index(last)

    new_index = index + direction
    if new_index < 0 or new_index >= len(items):
        return "break"

    new_item = items[new_index]

    # yeni satırı DAHİL ET
    device_tree.selection_add(new_item)
    device_tree.focus(new_item)
    device_tree.see(new_item)

    return "break"

def select_all_rows(event=None):
    items = device_tree.get_children()
    if not items:
        return "break"

    device_tree.selection_set(items)
    device_tree.focus(items[0])
    device_tree.see(items[0])
    return "break"

    
def get_selected_devices():
    selected = []
    for item in device_tree.selection():
        values = device_tree.item(item)["values"]
        if not values:
            continue

        ip = values[1]
        dev = next((d for d in devices if d["ip"] == ip), None)
        if dev:
            selected.append(dev)

    return selected

def on_tree_arrow(direction):
    move_selection(direction)
    return "break"   # 🔴 EN KRİTİK SATIR
    
def move_selection(direction):
    items = device_tree.get_children()
    if not items:
        return

    sel = device_tree.selection()
    if sel:
        index = items.index(sel[0])
    else:
        index = 0

    new_index = index + direction

    if new_index < 0 or new_index >= len(items):
        return

    device_tree.selection_set(items[new_index])
    device_tree.focus(items[new_index])
    device_tree.see(items[new_index])
    def write_ip_from_selection():
        global started_from_entry

        # 🔴 BİRDEN FAZLA SEÇİM VARSA HİÇBİR ŞEY YAPMA
        if len(device_tree.selection()) > 1:
            return

def get_paged_devices(filtered_devices):
    global total_pages

    total_pages = max(1, (len(filtered_devices) + PAGE_SIZE - 1) // PAGE_SIZE)

    start = (current_page - 1) * PAGE_SIZE
    end = start + PAGE_SIZE

    return filtered_devices[start:end]

def refresh_device_list(keep_selection=False):
    # mevcut seçimi hatırla
    selected_ip = None
    if keep_selection:
        sel = device_tree.selection()
        if sel:
            selected_ip = device_tree.item(sel[0])["values"][1]

    device_tree.delete(*device_tree.get_children())
    filtered = [d for d in devices if device_matches_filters(d)]
    paged_devices = get_paged_devices(filtered)

    for d in paged_devices:
        if not device_matches_filters(d):
            continue

        latency_txt = "-" if d.get("latency") is None else f"{d['latency']:.1f}"
        device_tree.insert(
            "",
            tk.END,
            values=(
                d.get("device", ""),
                d.get("ip", ""),
                latency_txt,
                d.get("last_ping") or "-",
                d.get("name", ""),     # 👈 DEVICE NAME
                d.get("model", ""),
                d.get("mac", ""),
                d.get("location", ""),
                d.get("unit", ""),
                d.get("description", "")
            ),
            tags=(d.get("status", "UNKNOWN"),)
        )

    # aynı IP varsa onu tekrar seç
    if selected_ip:
        for item in device_tree.get_children():
            if device_tree.item(item)["values"][1] == selected_ip:
                device_tree.selection_set(item)
                device_tree.focus(item)
                device_tree.see(item)
                write_ip_from_selection()
                return

    # yoksa ilk satırı seç
    items = device_tree.get_children()
    if items:
        device_tree.selection_set(items[0])
        device_tree.focus(items[0])
        device_tree.see(items[0])
        write_ip_from_selection()

    update_page_label()
    device_tree.update_idletasks()
    


def write_ip_from_selection():
    global started_from_entry

    if started_from_entry:
        return

    sel = device_tree.selection()
    if not sel:
        return

    ip_entry.delete(0, tk.END)
    ip_entry.insert(0, device_tree.item(sel[0])["values"][1])

def on_tree_select(event=None):
    global started_from_entry
    started_from_entry = False   # 👈 kilidi burada açıyoruz
    write_ip_from_selection()

def on_double_click(event):
    global started_from_entry
    row_id = device_tree.identify_row(event.y)
    if not row_id:
        return

    device_tree.selection_set(row_id)
    device_tree.focus(row_id)
    write_ip_from_selection()
    started_from_entry = False

    root.after(50, start_ping)

def move_focus_horizontal(direction):
    current = root.focus_get()

    if current not in top_controls:
        top_controls[0].focus_set()
        return

    index = top_controls.index(current)
    new_index = index + direction

    if 0 <= new_index < len(top_controls):
        top_controls[new_index].focus_set()

    refresh_device_list(keep_selection=True)

def show_sort_menu(event, col):
    menu = tk.Menu(root, tearoff=0)

    menu.add_command(
        label="A'dan Z'ye Sırala",
        command=lambda: sort_devices_by_column(col, reverse=False)
    )

    menu.add_command(
        label="Z'den A'ya Sırala",
        command=lambda: sort_devices_by_column(col, reverse=True)
    )

    menu.tk_popup(event.x_root, event.y_root)



# ---------------- CONTEXT MENU ----------------
def start_bulk_ping():
    if not can_start_new_task():
        return
    lock_ui("BULK")
    global is_bulk_running, bulk_total, bulk_done

    if is_running:
        messagebox.showwarning(
            "Uyarı",
            "Tekli ping çalışırken toplu ping başlatılamaz."
        )
        return

    if is_bulk_running:
        return

    # ✅ SADECE SEÇİLENLER
    devices_to_ping = get_selected_devices()

    if not devices_to_ping:
        messagebox.showinfo(
            "Bilgi",
            "Lütfen en az bir cihaz seçin"
        )
        return

    is_bulk_running = True

    bulk_total = len(devices_to_ping)
    bulk_done = 0

    bulk_status_label.config(
        text=f"Toplu Ping: 0 / {bulk_total}"
    )

    start_btn.ui_set_enabled(False)
    refresh_btn.ui_set_enabled(False)
    add_btn.ui_set_enabled(False)
    start_btn.config(text="⏹ Durdur")
    start_btn.ui_set_enabled(True)
    threading.Thread(
        target=bulk_ping_worker,
        args=(devices_to_ping,),
        daemon=True
    ).start()

def start_bulk_ping_all_filtered():
    global is_bulk_running, bulk_total, bulk_done

    if is_running:
        messagebox.showwarning(
            "Uyarı",
            "Tekli ping çalışırken toplu ping başlatılamaz."
        )
        return

    if is_bulk_running:
        return

    # 🔴 SAYFAYA DEĞİL → TÜM FİLTRELİ CİHAZLAR
    filtered_devices = [
        d for d in devices
        if device_matches_filters(d)
    ]

    if not filtered_devices:
        messagebox.showinfo(
            "Bilgi",
            "Filtrelere uyan cihaz bulunamadı"
        )
        return

    # 🔥 KRİTİK SATIRLAR (SENDEN EKSİK OLANLAR)
    bulk_total = len(filtered_devices)
    bulk_done = 0

    bulk_status_label.config(
        text=f"Toplu Ping: 0 / {bulk_total}"
    )

    is_bulk_running = True

    # 🔒 UI kilidi
    start_btn.ui_set_enabled(False)
    refresh_btn.ui_set_enabled(False)
    add_btn.ui_set_enabled(False)

    threading.Thread(
        target=bulk_ping_worker,
        args=(filtered_devices,),
        daemon=True
    ).start()


def open_add_device_window():
    ip = ip_entry.get().strip()

    if not ip:
        messagebox.showwarning("Uyarı", "Önce IP adresi giriniz")
        return

    if ip_exists(ip):
        messagebox.showwarning("IP Çakışması", f"Bu IP zaten kayıtlı:\n{ip}")
        return

    win = tk.Toplevel(root)
    win.transient(root)   # 🔴 KRİTİK
    win.grab_set()
    win.focus_force()
    win.title("Yeni Cihaz Ekle")
    win.resizable(False, False)
    win.grab_set()

    fields = [
        ("Device Name", ""),
        ("IP Address", ip),
        ("Device", ""),
        ("Model", ""),
        ("MAC", ""),
        ("Location", ""),
        ("Unit", ""),
        ("Description", ""),
    ]

    entries = {}

    for i, (label, value) in enumerate(fields):
        tk.Label(win, text=label).grid(row=i, column=0, sticky="w", padx=10, pady=4)

        ent = tk.Entry(win, width=40)
        ent.grid(row=i, column=1, padx=10, pady=4)
        ent.insert(0, value)

        entries[label] = ent
    def save_new_device():
        new_ip = entries["IP Address"].get().strip()

        if not new_ip:
            messagebox.showwarning("Hata", "IP Address boş olamaz")
            return

        if ip_exists(new_ip):
            messagebox.showwarning(
                "IP Çakışması",
                f"Bu IP zaten kayıtlı:\n{new_ip}"
            )
            return

        new_device = {
            "name": entries["Device Name"].get(),
            "ip": new_ip,
            "device": entries["Device"].get(),
            "model": entries["Model"].get(),
            "mac": entries["MAC"].get(),
            "location": entries["Location"].get(),
            "unit": entries["Unit"].get(),
            "description": entries["Description"].get(),
            "latency": None,
            "last_ping": None,
            "status": "UNKNOWN"
        }

        # ✅ 1️⃣ Excel'e yaz
        from device_loader import add_device_to_excel
        add_device_to_excel(new_device, excel_path, excel_mapping)

        # ✅ 2️⃣ Excel'den TEKRAR OKU (tek gerçek kaynak)
        refresh_from_excel()

        win.destroy()

    btns = tk.Frame(win)
    btns.grid(row=len(fields), column=0, columnspan=2, pady=10)

    tk.Button(btns, text="Kaydet", width=12, command=save_new_device).pack(side=tk.LEFT, padx=5)
    tk.Button(btns, text="İptal", width=12, command=win.destroy).pack(side=tk.LEFT, padx=5)
        
    
        
def show_device_details():
    sel = device_tree.selection()
    if not sel:
        return

    item = device_tree.item(sel[0])
    ip = item["values"][1]

    device = next((d for d in devices if d["ip"] == ip), None)
    if not device:
        return

    win = tk.Toplevel(root)
    win.title("Cihaz Detayları")
    win.resizable(False, False)
    win.grab_set()  # modal pencere

    fields = [
        ("Device Name", device.get("name"), True),
        ("IP Address", device.get("ip"), True),
        ("Device", device.get("device"), True),
        ("Model", device.get("model"), True),
        ("MAC", device.get("mac"), True),
        ("Location", device.get("location"), True),
        ("Unit", device.get("unit"), True),
        ("Description", device.get("description"), True),
    ]

    entries = {}

    for i, (label, value, editable) in enumerate(fields):
        tk.Label(win, text=label).grid(row=i, column=0, sticky="w", padx=10, pady=4)

        ent = tk.Entry(win, width=40)
        ent.grid(row=i, column=1, padx=10, pady=4)
        ent.insert(0, value if value else "")

        if not editable:
            ent.config(state="disabled")

        entries[label] = ent
    def save_changes():
        new_ip = entries["IP Address"].get().strip()

        if not new_ip:
            tk.messagebox.showwarning("Hata", "IP Address boş olamaz")
            return

        if ip_exists(new_ip, exclude_device=device):
            tk.messagebox.showwarning(
                "IP Çakışması",
                f"Bu IP zaten başka bir cihaza ait:\n{new_ip}"
            )
            return

        old_ip = device["ip"]

        device["name"] = entries["Device Name"].get()
        device["ip"] = new_ip
        device["device"] = entries["Device"].get()
        device["model"] = entries["Model"].get()
        device["mac"] = entries["MAC"].get()
        device["location"] = entries["Location"].get()
        device["unit"] = entries["Unit"].get()
        device["description"] = entries["Description"].get()

        from device_loader import update_device_in_excel
        update_device_in_excel(
            old_ip,
            device,
            excel_path,
            excel_mapping
        )

        save_devices(devices)
        refresh_device_list(keep_selection=True)
        win.destroy()

    btns = tk.Frame(win)
    btns.grid(row=len(fields), column=0, columnspan=2, pady=10)

    tk.Button(btns, text="Kaydet", width=12, command=save_changes).pack(side=tk.LEFT, padx=5)
    tk.Button(btns, text="İptal", width=12, command=win.destroy).pack(side=tk.LEFT, padx=5)
def copy_selected_ip():
    sel = device_tree.selection()
    if not sel:
        return
    ip = device_tree.item(sel[0])["values"][1]
    root.clipboard_clear()
    root.clipboard_append(ip)

def delete_selected_device():
    sel = device_tree.selection()
    if not sel:
        return

    item = device_tree.item(sel[0])
    ip = item["values"][1]

    answer = messagebox.askyesno(
        "Cihaz Sil",
        f"{ip} adresli cihaz silinsin mi?\n\nBu işlem geri alınamaz."
    )

    if not answer:
        return

    # 1️⃣ Excel’den sil
    from device_loader import delete_device_from_excel
    delete_device_from_excel(ip, excel_path, excel_mapping)

    # 2️⃣ Excel’den yeniden yükle
    refresh_from_excel()


def open_filter_window(field):
    win = tk.Toplevel(root)
    win.title(f"{FILTERABLE_FIELDS[field]} Filtre")
    win.geometry("320x450")
    win.resizable(False, False)
    win.grab_set()

    # ================== ARAMA ==================
    search_var = tk.StringVar()
    search_entry = tk.Entry(win, textvariable=search_var)
    search_entry.pack(fill=tk.X, padx=10, pady=(10, 5))

    # ================== ÜST BUTONLAR ==================
    top_btns = tk.Frame(win)
    top_btns.pack(fill=tk.X, padx=10, pady=5)

    def select_all():
        pass

    def clear_all():
        pass

    tk.Button(top_btns, text="☑️ Tümünü Seç", command=lambda: select_all()).pack(side=tk.LEFT)
    tk.Button(top_btns, text="❌ Temizle", command=lambda: clear_all()).pack(side=tk.LEFT, padx=5)

    # ================== SCROLLABLE ALAN ==================
    list_container = tk.Frame(win)
    list_container.pack(fill=tk.BOTH, expand=True, padx=10)

    canvas = tk.Canvas(list_container, borderwidth=0, highlightthickness=0)
    scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=canvas.yview)

    scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(2, 0))
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    canvas.configure(yscrollcommand=scrollbar.set)

    scroll_frame = tk.Frame(canvas)
    window_id = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")

    # ✅ Canvas genişliğini iç frame'e uydur
    def _resize_canvas(event):
        canvas.itemconfig(window_id, width=event.width)

    canvas.bind("<Configure>", _resize_canvas)

    # ✅ scrollregion + scrollbar thumb güncelle
    def _update_scrollregion(event=None):
        canvas.configure(scrollregion=canvas.bbox("all"))
        scrollbar.set(*canvas.yview())

    scroll_frame.bind("<Configure>", _update_scrollregion)

    # ✅ Windows mousewheel scroll (Checkbox üstünde bile çalışır)
    def _on_mousewheel(event):
        if not canvas.winfo_exists():
            return "break"

        delta = int(event.delta / 120)
        if delta == 0:
            delta = 1 if event.delta > 0 else -1

        canvas.yview_scroll(-delta, "units")
        return "break"

    # 🔥 Mousewheel'i pencereye bağla ki her yerde çalışsın
    win.bind_all("<MouseWheel>", _on_mousewheel)

    # pencere kapanınca bind_all temizle (yoksa tüm uygulama etkilenir)
    def _cleanup_bindings():
        try:
            win.unbind_all("<MouseWheel>")
        except Exception:
            pass

    win.protocol("WM_DELETE_WINDOW", lambda: ( _cleanup_bindings(), win.destroy() ))

    # ================== ALT OK BUTONU ==================
    bottom = tk.Frame(win)
    bottom.pack(fill=tk.X, pady=10)

    def apply_filters():
        global current_page
        current_page = 1
        active_filters[field].clear()

        for val, var in vars_map.items():
            if var.get():
                active_filters[field].add(val)

        refresh_device_list()
        update_column_headers()
        _cleanup_bindings()
        win.destroy()

    tk.Button(bottom, text="OK", width=10, command=apply_filters).pack(side=tk.RIGHT, padx=10)

    # ================== VERİLER ==================
    def ip_key(ip):
        try:
            return tuple(int(p) for p in ip.split("."))
        except:
            return (999, 999, 999, 999)

    values = sorted(
        set(str(d.get(field)) for d in devices if d.get(field)),
        key=ip_key if field == "ip" else lambda x: x.lower()
    )
    vars_map = {}
    checkbuttons = {}

    def render_list():
        # eski checkboxları sil
        for chk in checkbuttons.values():
            chk.destroy()

        checkbuttons.clear()
        vars_map.clear()

        keyword = search_var.get().lower()

        for val in values:
            if keyword and keyword not in val.lower():
                continue

            var = tk.BooleanVar(value=val in active_filters[field])
            chk = tk.Checkbutton(scroll_frame, text=val, variable=var)
            chk.pack(anchor="w")

            vars_map[val] = var
            checkbuttons[val] = chk

        win.update_idletasks()
        _update_scrollregion()

    def select_all():
        for var in vars_map.values():
            var.set(True)

    def clear_all():
        for var in vars_map.values():
            var.set(False)

    search_var.trace_add("write", lambda *args: render_list())
    render_list()
    
def clear_all_filters():
    global current_page

    # 1️⃣ tüm filtre setlerini boşalt
    for field in active_filters:
        active_filters[field].clear()

    # 2️⃣ kolon başlıklarını eski haline döndür
    for col in COLUMN_TO_FIELD:
        device_tree.heading(col, text=f"{col} ▼")

    # 3️⃣ sayfayı başa al
    current_page = 1

    # 4️⃣ listeyi yenile
    refresh_device_list()
    update_column_headers()

def clear_single_filter(field):
    global current_page

    active_filters[field].clear()
    current_page = 1
    refresh_device_list()
    update_column_headers()

def show_column_menu(event, col):
    menu = tk.Menu(root, tearoff=0)

    menu.add_command(
        label="A'dan Z'ye Sırala",
        command=lambda: sort_devices_by_column(col, reverse=False)
    )
    menu.add_command(
        label="Z'den A'ya Sırala",
        command=lambda: sort_devices_by_column(col, reverse=True)
    )

    if col in COLUMN_TO_FIELD:
        menu.add_separator()
        menu.add_command(
            label="Filtrele",
            command=lambda: open_filter_window(COLUMN_TO_FIELD[col])
        )
        menu.add_command(
            label="Filtreyi Temizle",
            command=lambda: clear_single_filter(COLUMN_TO_FIELD[col])
        )
   

    menu.tk_popup(event.x_root, event.y_root)
    menu.grab_release()

def show_context_menu(event):
    # Sağ tık yapılan satırı bul
    row_id = device_tree.identify_row(event.y)

    if not row_id:
        return

    # Eğer o satır zaten seçiliyse -> HİÇBİR ŞEY YAPMA
    if row_id not in device_tree.selection():
        # değilse sadece o satırı seç (tekli senaryo)
        device_tree.selection_set(row_id)
        device_tree.focus(row_id)
        write_ip_from_selection()

    context_menu.tk_popup(event.x_root, event.y_root)

# ---------------- UI ----------------
root = tk.Tk()
root.title("Ping Monitor")
root.iconbitmap(resource_path(os.path.join("assets", "app.ico")))
root.geometry("1100x650")
root.minsize(1100, 650)
root.configure(bg=BG_COLOR)

# Modern UI assets (rounded controls + background)
_ui_assets = load_ui_assets()

# ---------------- BACKGROUND ----------------
bg_photo = _ui_assets["bg"]
bg_label = tk.Label(root, image=bg_photo, bg=BG_COLOR, bd=0, highlightthickness=0)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# ---------------- TTK DARK STYLE (Treeview/Scrollbars) ----------------
style = apply_ttk_dark_style(root)
style.configure("Treeview", font=(FONT_NAME, 11))
style.configure("Treeview.Heading", font=(FONT_NAME, 11, "bold"))

top = tk.Frame(root, bg=BG_COLOR)

top.pack(fill=tk.X, padx=10, pady=5)
left_controls = tk.Frame(top, bg=BG_COLOR)
left_controls.pack(side=tk.LEFT)

right_controls = tk.Frame(top, bg=BG_COLOR)
right_controls.pack(side=tk.RIGHT)

tk.Label(left_controls, text="IP:", font=(FONT_NAME, 11), bg=BG_COLOR, fg=FG_COLOR).pack(side=tk.LEFT)

ip_entry_container, ip_entry = make_rounded_entry(left_controls, _ui_assets["entry_bg"], font=(FONT_NAME, 11))
ip_entry_container.pack(side=tk.LEFT, padx=8)

start_btn = tk.Button(left_controls, text="▶ Başlat", command=toggle_ping)
style_rounded_button(start_btn, _ui_assets, wide=False)
start_btn.pack(side=tk.LEFT, padx=6)

refresh_btn = tk.Button(left_controls, text="⟳ Yenile", command=refresh_from_excel)
style_rounded_button(refresh_btn, _ui_assets, wide=False)
refresh_btn.pack(side=tk.LEFT, padx=6)

add_btn = tk.Button(left_controls, text="➕ Ekle", command=open_add_device_window)
style_rounded_button(add_btn, _ui_assets, wide=False)
add_btn.pack(side=tk.LEFT, padx=6)

excel_btn = tk.Button(left_controls, text="📂 Excel Seç", command=select_excel_file)
style_rounded_button(excel_btn, _ui_assets, wide=True)
excel_btn.pack(side=tk.LEFT, padx=6)

tk.Label(right_controls, text="Ara:", font=(FONT_NAME, 11), bg=BG_COLOR, fg=FG_COLOR).pack(side=tk.LEFT, padx=(0, 8))

search_container, search_entry = make_rounded_entry(right_controls, _ui_assets["search_bg"], font=(FONT_NAME, 11))
search_container.pack(side=tk.LEFT)


def on_search_change(event=None):
    global search_text
    search_text = search_entry.get().strip()
    refresh_device_list()

search_entry.bind("<KeyRelease>", on_search_change)


top_controls = [
    ip_entry,
    start_btn,
    refresh_btn,
    add_btn
]

main = tk.PanedWindow(root, orient=tk.HORIZONTAL, bg=BG_COLOR, bd=0, sashwidth=6, sashrelief='flat')
main.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

# ✅ OUTPUT BOX container (scrollbar + text)
output_container = tk.Frame(main, bg=BG_COLOR)
main.add(output_container)

output_scroll = ttk.Scrollbar(output_container, orient=tk.VERTICAL)
output_scroll.pack(side=tk.RIGHT, fill=tk.Y)

output_box = tk.Text(
    output_container,
    state=tk.DISABLED,
    font=(FONT_NAME, 11),
    bg=PANEL_COLOR,
    fg=FG_COLOR,
    insertbackground=FG_COLOR,
    relief='flat',
    bd=0,
    highlightthickness=0,
    yscrollcommand=output_scroll.set
)
output_box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

output_scroll.config(command=output_box.yview)

right = tk.Frame(main, bg=BG_COLOR)
main.add(right)


# 🔹 Treeview + Scrollbar için container
tree_container = tk.Frame(right, bg=BG_COLOR)
tree_container.pack(fill=tk.BOTH, expand=True)


# 🔹 Dikey scrollbar
tree_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL)
tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

# 🔹 Yatay scrollbar (DİKKAT: burada device_tree YOK)
tree_scroll_x = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL)
tree_scroll_x.pack(side=tk.BOTTOM, fill=tk.X)

# 🔹 Treeview

device_tree = ttk.Treeview(
    tree_container,
    columns=cols,
    show="headings",
    selectmode="extended",
    yscrollcommand=tree_scroll.set,
    xscrollcommand=tree_scroll_x.set,
    takefocus=False
)

def on_heading_click(event):
    region = device_tree.identify_region(event.x, event.y)
    if region != "heading":
        return

    col_id = device_tree.identify_column(event.x)
    col_index = int(col_id.replace("#", "")) - 1
    col_name = cols[col_index]

    show_column_menu(event, col_name)
    return "break"  

# 🔑 Treeview başlıklarını TANIMLA
for c in cols:
    device_tree.heading(c, text=c)
    device_tree.column(c, width=COLUMN_WIDTHS[c], anchor="w")

for c in cols:
    if c in COLUMN_TO_FIELD:
        device_tree.heading(c, text=f"{c} ▼")
    else:
        device_tree.heading(c, text=c)

# 🔹 Scrollbar ↔ Treeview bağlantısı
tree_scroll.config(command=device_tree.yview)
tree_scroll_x.config(command=device_tree.xview)

# 🔹 Column genişlikleri
for c in cols:
    device_tree.column(
        c,
        width=COLUMN_WIDTHS[c],
        minwidth=COLUMN_WIDTHS[c],
        stretch=False
    )

device_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
device_tree.update_idletasks()

bulk_status_label = tk.Label(right, text="", font=(FONT_NAME, 10), bg=BG_COLOR, fg=MUTED_FG)
bulk_status_label.pack(fill=tk.X, padx=10, pady=(0, 5))
# ---------------- PAGINATION UI ----------------
pagination = tk.Frame(right, bg=BG_COLOR)
pagination.pack(fill=tk.X, pady=5)

page_label = tk.Label(pagination, text="Sayfa 1 / 1", bg=BG_COLOR, fg=FG_COLOR)
page_label.pack(side=tk.LEFT, padx=10)

# ---------------- FOOTER / SIGNATURE ----------------
footer = tk.Frame(right, bg=BG_COLOR)
footer.pack(fill=tk.X, side=tk.BOTTOM, padx=8, pady=4)

signature_label = tk.Label(
    footer,
    text="Anur",
    font=(FONT_NAME, 9),
    bg=BG_COLOR,
    fg="#7f8c8d"   # gri – göz yormaz
)
signature_label.pack(side=tk.RIGHT)

def update_page_label():
    page_label.config(text=f"Sayfa {current_page} / {total_pages}")

def prev_page():
    global current_page
    if current_page > 1:
        current_page -= 1
        refresh_device_list()
        update_page_label()

def next_page():
    global current_page
    if current_page < total_pages:
        current_page += 1
        refresh_device_list()
        update_page_label()

prev_btn = tk.Button(pagination, text="◀ Önceki", command=prev_page)
style_rounded_button(prev_btn, _ui_assets, wide=False)
prev_btn.pack(side=tk.LEFT, padx=(6, 6))

next_btn = tk.Button(pagination, text="Sonraki ▶", command=next_page)
style_rounded_button(next_btn, _ui_assets, wide=False)
next_btn.pack(side=tk.LEFT, padx=(0, 6))

# BINDINGS
device_tree.bind("<Button-1>", on_heading_click)
device_tree.bind("<<TreeviewSelect>>", on_tree_select)
device_tree.bind("<Double-1>", on_double_click)
device_tree.bind("<Up>", lambda e: on_tree_arrow(-1))
device_tree.bind("<Down>", lambda e: on_tree_arrow(1))
device_tree.bind("<Shift-Up>", lambda e: extend_selection(-1))
device_tree.bind("<Shift-Down>", lambda e: extend_selection(1))
device_tree.bind("<Control-Up>", lambda e: extend_selection(-1))
device_tree.bind("<Control-Down>", lambda e: extend_selection(1))
device_tree.bind("<Control-a>", select_all_rows)
device_tree.bind("<Command-a>", select_all_rows)  # macOS için

def on_mousewheel(event):
    device_tree.yview_scroll(int(-1*(event.delta/120)), "units")

device_tree.bind("<MouseWheel>", on_mousewheel)        # Windows
device_tree.bind("<Button-4>", lambda e: device_tree.yview_scroll(-1, "units"))  # Mac
device_tree.bind("<Button-5>", lambda e: device_tree.yview_scroll(1, "units"))   # Mac


device_tree.bind("<Button-3>", show_context_menu)
device_tree.bind("<Button-2>", show_context_menu)
def ctrl_click_select(event):
    row = device_tree.identify_row(event.y)
    if not row:
        return "break"

    if row in device_tree.selection():
        device_tree.selection_remove(row)
    else:
        device_tree.selection_add(row)

    device_tree.focus(row)
    return "break"

device_tree.bind("<Control-Button-1>", ctrl_click_select)
root.bind("<Shift-F10>", show_context_menu)


def safe_start_ping(event=None):
    if is_bulk_running:
        return
    if len(device_tree.selection()) > 1:
        return
    start_ping(event)

root.bind("<Return>", safe_start_ping)
root.bind("<Escape>", stop_ping)
root.bind("<Left>", lambda e: move_focus_horizontal(-1))
root.bind("<Right>", lambda e: move_focus_horizontal(1))

# RENKLER
device_tree.tag_configure("UNKNOWN", foreground="#7f8c8d")
device_tree.tag_configure("FAST", foreground="#1e8449")
device_tree.tag_configure("NORMAL", foreground="#27ae60")
device_tree.tag_configure("SLOW", foreground="#b7950b")
device_tree.tag_configure("VERY_SLOW", foreground="#ca6f1e")
device_tree.tag_configure("DOWN", foreground="#c0392b")

# CONTEXT MENU
context_menu = tk.Menu(root, tearoff=0)
context_menu.add_separator()
context_menu.add_command(
    label="🗑 Cihazı Sil",
    command=delete_selected_device
)
context_menu.add_command(label="▶ Ping Başlat", command=start_ping_from_menu)
context_menu.add_command(label="⏹ Ping Durdur", command=stop_ping)
context_menu.add_command(label="🧭 Traceroute", command=start_traceroute_selected)
context_menu.add_command(
    label="🔌 Port Test (Hızlı: Öncelikli + Diğer)",
    command=lambda: start_port_test_selected("fast")
)

context_menu.add_command(
    label="🔥 Port Test (FULL 1-65535) [Riskli/Uzun]",
    command=lambda: start_port_test_selected("full")
)
context_menu.add_command(label="🌐 NSLOOKUP (DNS Test)", command=start_nslookup_selected)
context_menu.add_separator()
context_menu.add_command(
    label="📡 Seçilenlere Toplu Ping",
    command=start_bulk_ping
)
context_menu.add_command(
    label="🌐 Filtrelenmiş TÜM Cihazlara Ping",
    command=start_bulk_ping_all_filtered
)
context_menu.add_separator()

# 🔴 YENİ EKLENEN
context_menu.add_command(label="📝 Cihaz Detayları", command=show_device_details)

context_menu.add_separator()
context_menu.add_command(label="📋 IP Kopyala", command=copy_selected_ip)
context_menu.add_separator()
context_menu.add_command(label="🔄 Excel'den Yenile", command=refresh_from_excel)
context_menu.add_separator()
context_menu.add_command(
    label="🧹 Tüm Filtreleri Temizle",
    command=clear_all_filters
)

# ---------------- START ----------------
load_config()

devices = []

cached_devices = load_devices()

if excel_path and excel_mapping:
    excel_devices = load_devices_from_excel(excel_path, excel_mapping)

    for ex in excel_devices:
        cached = next((d for d in cached_devices if d.get("ip") == ex.get("ip")), None)

        if cached:
            ex["latency"] = cached.get("latency")
            ex["last_ping"] = cached.get("last_ping")
            ex["status"] = cached.get("status", "UNKNOWN")
        else:
            ex["latency"] = None
            ex["last_ping"] = None
            ex["status"] = "UNKNOWN"

        devices.append(ex)

refresh_device_list()
root.after(100, process_ui_queue)
root.mainloop()
